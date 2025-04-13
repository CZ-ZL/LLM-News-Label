import pandas as pd
import argparse
from sklearn.metrics import confusion_matrix, precision_score, recall_score, accuracy_score
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def write_dataframe_to_sheet(ws, df, start_row=1, start_col=1):
    """
    Writes the content of a DataFrame to an openpyxl worksheet starting at the given row/column.
    Uses dataframe_to_rows so that index and header are included.
    """
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

def main():
    # Set up the argument parser.
    parser = argparse.ArgumentParser(
        description='Append confusion matrices and metrics (merged) into the original Excel file, keeping all original sheets intact.'
    )
    parser.add_argument('--input_file', type=str, required=True,
                        help='Path to the original input Excel file (will be preserved)')
    parser.add_argument('--data_sheet', type=str, required=True,
                        help='Name of the sheet containing the data')
    parser.add_argument('--output_file', type=str, default='output_merged.xlsx',
                        help='Path for the output Excel file (a copy with added results)')
    args = parser.parse_args()

    input_file = args.input_file
    data_sheet = args.data_sheet
    output_file = args.output_file

    # Load the data using pandas.
    try:
        df = pd.read_excel(input_file, sheet_name=data_sheet)
    except Exception as e:
        print(f"Error reading Excel file/sheet: {e}")
        return

    # Ensure required columns exist.
    required_columns = ['Market Direction', 'LLM Prompt Label', 'Expert Label', 'Market Label']
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        print(f"Error: The following required columns are missing: {', '.join(missing)}")
        return

    true_labels = df['Market Direction']
    predicted_columns = ['LLM Prompt Label', 'Expert Label', 'Market Label']
    class_labels = [1, -1]

    # Compute confusion matrices and metrics for each predicted column.
    matrices = {}
    metrics_dict = {}
    for col in predicted_columns:
        predicted = df[col]
        cm = confusion_matrix(true_labels, predicted, labels=class_labels)
        cm_df = pd.DataFrame(
            cm,
            index=[f"True {i}" for i in class_labels],
            columns=[f"Pred {i}" for i in class_labels]
        )
        matrices[col] = cm_df

        precision = precision_score(true_labels, predicted, average='weighted', zero_division=0)
        recall_val = recall_score(true_labels, predicted, average='weighted', zero_division=0)
        accuracy = accuracy_score(true_labels, predicted)
        metrics_df = pd.DataFrame({
            'Metric': ['Precision', 'Recall', 'Accuracy'],
            'Score': [precision, recall_val, accuracy]
        })
        metrics_dict[col] = metrics_df

        print(f"Computed confusion matrix and metrics for column: '{col}'.")

    # Load the original workbook so that all original sheets remain.
    wb = load_workbook(input_file)

    # For each predicted column, create new sheets (temporary and merged).
    for col in predicted_columns:
        # Define sheet names.
        matrix_sheet_name = f"Matrix_{col.replace(' ', '_')}"
        metrics_sheet_name = f"Metrics_{col.replace(' ', '_')}"
        merged_sheet_name = f"Merged_{col.replace(' ', '_')}"

        # Create and write the matrix sheet.
        matrix_ws = wb.create_sheet(title=matrix_sheet_name)
        matrix_ws.cell(row=1, column=1, value=f"Confusion Matrix for {col}")
        write_dataframe_to_sheet(matrix_ws, matrices[col], start_row=3, start_col=1)

        # Create and write the metrics sheet.
        metrics_ws = wb.create_sheet(title=metrics_sheet_name)
        metrics_ws.cell(row=1, column=1, value=f"Metrics for {col}")
        write_dataframe_to_sheet(metrics_ws, metrics_dict[col], start_row=3, start_col=1)

        # Create the merged sheet.
        merged_ws = wb.create_sheet(title=merged_sheet_name)
        # Copy contents from the matrix sheet.
        for r_idx, row in enumerate(matrix_ws.iter_rows(values_only=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                merged_ws.cell(row=r_idx, column=c_idx, value=value)
        matrix_rows = matrix_ws.max_row
        gap = 2  # leave two blank rows between sections
        metrics_start_row = matrix_rows + gap
        # Copy contents from the metrics sheet.
        for r_idx, row in enumerate(metrics_ws.iter_rows(values_only=True), start=metrics_start_row):
            for c_idx, value in enumerate(row, start=1):
                merged_ws.cell(row=r_idx, column=c_idx, value=value)

        print(f"Created merged sheet '{merged_sheet_name}' for predicted column '{col}'.")

    # Save the workbook as a new file so the original file is preserved.
    wb.save(output_file)
    print(f"All results have been merged and saved to '{output_file}'. The original file remains unchanged.")

if __name__ == '__main__':
    main()
